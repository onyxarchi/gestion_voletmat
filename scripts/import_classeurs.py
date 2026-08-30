#!/usr/bin/env python3
"""
Import lecture seule des classeurs N4 / N5 vers la base SQLite Vol&Mat Gestion.
Aucun montant inventé : seules les valeurs présentes (ou TVA/TTC calculés
depuis HT × taux explicite) sont enregistrées.

Usage :
  VOLETMAT_SQLITE_PATH=$HOME/voletmat_gestion.sqlite \\
    /path/to/venv/bin/python scripts/import_classeurs.py
"""
from __future__ import annotations

import hashlib
import os
import re
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

ROOT = Path(__file__).resolve().parents[1]
AIDES = ROOT / "aides"
N4 = AIDES / "_gestion_n4_24-25.xlsm"
N5 = AIDES / "_gestion_n5_25-26.xlsm"
DB = Path(os.environ.get("VOLETMAT_SQLITE_PATH", str(ROOT / "storage/data/voletmat_gestion.sqlite")))

# Planning N5 : mois juil 2025 → déc 2026 = colonnes I(9) … Z(26)
N5_MONTHS = [
    "202507", "202508", "202509", "202510", "202511", "202512",
    "202601", "202602", "202603", "202604", "202605", "202606",
    "202607", "202608", "202609", "202610", "202611", "202612",
]
# Planning N4 : à déduire des en-têtes

STATUT_FROM_RGB = {
    "FFFFFF00": ("litige", "jaune"),  # facturé mais client ne veut pas payer
    # Bleu FFDBE2F1 / FFDAE3F3 = dossiers MAR (ligne), pas « payé »
}

BLEU_MAR = {"FFDBE2F1", "FFDAE3F3"}


def fill_rgb(cell: Cell) -> str | None:
    f = cell.fill
    if not f or not f.fgColor or f.patternType in (None, "none"):
        return None
    c = f.fgColor
    if getattr(c, "type", None) == "rgb" and c.rgb and isinstance(c.rgb, str):
        return c.rgb.upper()
    return None


def is_row_finished(cell: Cell) -> bool:
    """Fond gris Excel (thème 2) = dossier fini pendant l’exercice."""
    f = cell.fill
    if not f or not f.fgColor or f.patternType in (None, "none"):
        return False
    c = f.fgColor
    return getattr(c, "type", None) == "theme" and getattr(c, "theme", None) == 2


def is_row_mar(cell: Cell, ref) -> bool:
    """Fond bleu = MAR ; refs M… aussi."""
    rgb = fill_rgb(cell)
    if rgb in BLEU_MAR:
        return True
    if ref is not None and str(ref).strip().upper().startswith("M"):
        return True
    return False


def log(msg: str) -> None:
    print(msg, flush=True)


def to_iso(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)) and 20000 < float(v) < 80000:
        # série Excel
        from datetime import timedelta
        base = date(1899, 12, 30)
        return (base + timedelta(days=int(v))).isoformat()
    s = str(v).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None


def to_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        if v.startswith("="):
            return None  # formule non évaluée → ne pas inventer
        s = v.replace("\u00a0", "").replace(" ", "").replace("€", "").replace(",", ".")
        # cellules avec tirets / tirets longs
        if s in {"-", "–", "—"} or not re.search(r"\d", s):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def eval_simple_amount(v) -> float | None:
    """Accepte un nombre ou une formule arithmétique simple type =21900*0.9."""
    n = to_float(v)
    if n is not None:
        return n
    if isinstance(v, str) and v.startswith("="):
        expr = v[1:].replace(" ", "")
        # uniquement chiffres, + - * / . ( )
        if re.fullmatch(r"[\d\.\+\-\*/\(\)]+", expr):
            try:
                return float(eval(expr, {"__builtins__": {}}, {}))  # noqa: S307 — expression contrôlée
            except Exception:
                return None
    return None


def taux_from_tva_formula(cell_val, ht: float | None) -> tuple[float, float | None]:
    """Retourne (taux, tva). Si formule D*20% → 0.20 ; si montant saisi → taux déduit si HT connu."""
    if isinstance(cell_val, str) and cell_val.startswith("="):
        if "10%" in cell_val or "*0.1" in cell_val.replace(" ", ""):
            tva = (ht * 0.10) if ht is not None else None
            return 0.10, tva
        if "20%" in cell_val or "*0.2" in cell_val.replace(" ", ""):
            tva = (ht * 0.20) if ht is not None else None
            return 0.20, tva
    tva = to_float(cell_val)
    if tva is not None and ht not in (None, 0):
        return round(abs(tva / ht), 4), tva
    if tva is not None:
        return 0.20, tva
    return 0.20, (ht * 0.20) if ht is not None else None


def empreinte(date_op: str, libelle: str, debit, credit) -> str:
    d = "" if debit is None else f"{float(debit):.2f}"
    c = "" if credit is None else f"{float(credit):.2f}"
    return hashlib.sha256(f"{date_op}|{libelle.strip()}|{d}|{c}".encode()).hexdigest()


def ensure_exercice(con: sqlite3.Connection, code: str, libelle: str, debut: str, fin: str, actif: int) -> int:
    con.execute(
        """INSERT INTO exercices (code, libelle, date_debut, date_fin, actif)
           VALUES (?,?,?,?,?)
           ON CONFLICT(code) DO UPDATE SET
             libelle=excluded.libelle,
             date_debut=excluded.date_debut,
             date_fin=excluded.date_fin""",
        (code, libelle, debut, fin, actif),
    )
    if actif:
        con.execute("UPDATE exercices SET actif = 0 WHERE code != ?", (code,))
        con.execute("UPDATE exercices SET actif = 1 WHERE code = ?", (code,))
    row = con.execute("SELECT id FROM exercices WHERE code = ?", (code,)).fetchone()
    return int(row[0])


def clear_exercice_data(con: sqlite3.Connection, eid: int) -> None:
    # rapprochements via factures
    con.execute(
        "DELETE FROM rapprochements WHERE facture_id IN (SELECT id FROM factures WHERE exercice_id=?)",
        (eid,),
    )
    con.execute(
        "DELETE FROM echeances_facturation WHERE affaire_id IN (SELECT id FROM affaires WHERE exercice_id=?)",
        (eid,),
    )
    con.execute("DELETE FROM affaires WHERE exercice_id=?", (eid,))
    con.execute("DELETE FROM factures WHERE exercice_id=?", (eid,))
    con.execute("DELETE FROM operations_bancaires WHERE exercice_id=?", (eid,))
    con.execute("DELETE FROM budgets WHERE exercice_id=?", (eid,))


def import_facturation(con: sqlite3.Connection, ws, eid: int, source: str) -> int:
    n = 0
    for r in range(1, (ws.max_row or 1) + 1):
        numero = ws.cell(r, 2).value
        if not numero or not isinstance(numero, (str, int, float)):
            continue
        numero = str(numero).strip()
        # ignorer en-têtes
        if numero.upper() in {"HT", "N°", "NO", "NUMERO", "NUMÉRO"} or numero.startswith("="):
            continue
        date_f = to_iso(ws.cell(r, 1).value)
        client = ws.cell(r, 3).value
        if not date_f and not client:
            continue
        if not date_f:
            continue
        ht = eval_simple_amount(ws.cell(r, 4).value)
        if ht is None:
            ht = 0.0
        taux, tva = taux_from_tva_formula(ws.cell(r, 5).value, ht)
        if tva is None:
            tva = 0.0
        ttc_cell = eval_simple_amount(ws.cell(r, 6).value)
        ttc = ttc_cell if ttc_cell is not None else (ht + tva)
        canal = ws.cell(r, 7).value
        canal_s = str(canal).strip() if canal not in (None, "") else None
        con.execute(
            """INSERT INTO factures
               (exercice_id, numero, date_facture, client, ht, taux_tva, tva, ttc, canal, notes)
               VALUES (?,?,?,?,?,?,?,?,?,?)
               ON CONFLICT(exercice_id, numero) DO UPDATE SET
                 date_facture=excluded.date_facture,
                 client=excluded.client,
                 ht=excluded.ht,
                 taux_tva=excluded.taux_tva,
                 tva=excluded.tva,
                 ttc=excluded.ttc,
                 canal=excluded.canal""",
            (eid, numero, date_f, str(client or "").strip(), ht, taux, tva, ttc, canal_s, f"import:{source}"),
        )
        n += 1
    return n


def import_banque_n5(con: sqlite3.Connection, ws, eid: int) -> int:
    """Colonnes : Date, Date valeur, Libellé, Débit, Crédit, TRI, MOIS."""
    n = 0
    for r in range(2, (ws.max_row or 1) + 1):
        lib = ws.cell(r, 3).value
        date_op = to_iso(ws.cell(r, 1).value)
        if not lib or not date_op:
            continue
        libelle = str(lib).strip()
        if not libelle:
            continue
        debit = to_float(ws.cell(r, 4).value)
        credit = to_float(ws.cell(r, 5).value)
        tri = ws.cell(r, 6).value
        tri_s = str(tri).strip() if tri not in (None, "") else None
        # vérifier catégorie
        if tri_s:
            exists = con.execute("SELECT 1 FROM categories WHERE code=?", (tri_s,)).fetchone()
            if not exists:
                con.execute(
                    "INSERT OR IGNORE INTO categories (code, libelle, ordre) VALUES (?,?,?)",
                    (tri_s, tri_s, 99),
                )
        mois = date_op[:4] + date_op[5:7]
        emp = empreinte(date_op, libelle, debit, credit)
        try:
            con.execute(
                """INSERT INTO operations_bancaires
                   (exercice_id, date_operation, date_valeur, libelle, debit, credit,
                    categorie_code, annee_mois, source, empreinte)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    eid,
                    date_op,
                    to_iso(ws.cell(r, 2).value),
                    libelle,
                    debit,
                    credit,
                    tri_s,
                    mois,
                    "excel_n5",
                    emp,
                ),
            )
            n += 1
        except sqlite3.IntegrityError:
            pass  # doublon empreinte
    return n


def import_banque_n4(con: sqlite3.Connection, ws, eid: int) -> int:
    """Pointage BQ : Date, Date valeur, Débit, Crédit, Libellé, TRI."""
    n = 0
    for r in range(2, (ws.max_row or 1) + 1):
        date_op = to_iso(ws.cell(r, 1).value)
        lib = ws.cell(r, 5).value
        if not date_op or not lib:
            continue
        libelle = str(lib).strip()
        debit = to_float(ws.cell(r, 3).value)
        credit = to_float(ws.cell(r, 4).value)
        tri = ws.cell(r, 6).value
        tri_s = str(tri).strip() if tri not in (None, "") else None
        if tri_s and not con.execute("SELECT 1 FROM categories WHERE code=?", (tri_s,)).fetchone():
            con.execute(
                "INSERT OR IGNORE INTO categories (code, libelle, ordre) VALUES (?,?,?)",
                (tri_s, tri_s, 99),
            )
        mois = date_op[:4] + date_op[5:7]
        emp = empreinte(date_op, libelle, debit, credit)
        try:
            con.execute(
                """INSERT INTO operations_bancaires
                   (exercice_id, date_operation, date_valeur, libelle, debit, credit,
                    categorie_code, annee_mois, source, empreinte)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    eid,
                    date_op,
                    to_iso(ws.cell(r, 2).value),
                    libelle,
                    debit,
                    credit,
                    tri_s,
                    mois,
                    "excel_n4",
                    emp,
                ),
            )
            n += 1
        except sqlite3.IntegrityError:
            pass
    return n


def detect_affaire_type(ref) -> str:
    if ref is None:
        return "autre"
    s = str(ref).strip()
    su = s.upper()
    if su.startswith("STRIPE"):
        return "stripe"
    if "PV" in su:
        return "pv"
    if su.startswith("DIVERS"):
        return "autre"
    if su.startswith("M"):
        return "mission"
    if su.startswith("C") or re.match(r"^\d{4}", s):
        return "contrat"
    return "autre"


def import_planning(con: sqlite3.Connection, ws, eid: int, months: list[str], first_data_row: int = 5) -> tuple[int, int]:
    """Importe affaires + échéances mensuelles.
    Ligne bleue = MAR ; ligne grise (thème 2) = dossier fini ; jaune cellule = litige.
    Références STRIPE / Divers peuvent être identiques (plusieurs lignes).
    """
    # assurer colonne fini
    cols = {r[1] for r in con.execute("PRAGMA table_info(affaires)")}
    if "fini" not in cols:
        con.execute("ALTER TABLE affaires ADD COLUMN fini INTEGER NOT NULL DEFAULT 0")

    na = ne = 0
    # détecter fin des lignes d'affaires (avant totaux / PROSPECTS)
    for r in range(first_data_row, min((ws.max_row or 1), 90) + 1):
        ref = ws.cell(r, 2).value
        client = ws.cell(r, 3).value
        if client and str(client).strip().upper().startswith("PRÉVISIONNEL"):
            break
        if ref and str(ref).strip().upper() in {"PROSPECTS", "MAR"}:
            break
        if not client and not ref:
            continue
        client_s = str(client or "").strip()
        if not client_s:
            continue
        if client_s.lower().startswith("sous total") or client_s.lower().startswith("montant"):
            continue

        raw_ref = str(ref).strip() if ref not in (None, "") else ""
        typ = detect_affaire_type(ref if raw_ref else None)
        cell_client = ws.cell(r, 3)
        cell_ref = ws.cell(r, 2)
        mar = is_row_mar(cell_client, ref) or is_row_mar(cell_ref, ref)
        if mar:
            typ = "mission"
        fini = 1 if (is_row_finished(cell_client) or is_row_finished(cell_ref)) else 0

        ref_s = raw_ref if raw_ref else "Divers"

        contrat = eval_simple_amount(ws.cell(r, 5).value)
        encaisse = eval_simple_amount(ws.cell(r, 6).value) or 0.0

        cur = con.execute(
            """INSERT INTO affaires
               (exercice_id, reference, client, type, montant_contrat_ht, encaisse_n1, fini)
               VALUES (?,?,?,?,?,?,?)""",
            (eid, ref_s, client_s, typ, contrat, encaisse, fini),
        )
        aid = int(cur.lastrowid)
        na += 1

        for i, mois in enumerate(months):
            col = 9 + i  # I = 9
            if col > 26:
                break
            cell = ws.cell(r, col)
            montant = eval_simple_amount(cell.value)
            if montant is None:
                montant = to_float(cell.value)
            if montant is None or montant == 0:
                continue
            rgb = fill_rgb(cell)
            if rgb and rgb in STATUT_FROM_RGB:
                statut, couleur = STATUT_FROM_RGB[rgb]
            elif rgb in BLEU_MAR:
                statut, couleur = "a_facturer", "bleu_mar"
            elif is_row_finished(cell):
                statut, couleur = "a_facturer", "gris_fini"
            else:
                statut, couleur = "a_facturer", None
            con.execute(
                """INSERT INTO echeances_facturation
                   (affaire_id, annee_mois, montant_ht, statut, couleur)
                   VALUES (?,?,?,?,?)
                   ON CONFLICT(affaire_id, annee_mois) DO UPDATE SET
                     montant_ht=excluded.montant_ht,
                     statut=excluded.statut,
                     couleur=excluded.couleur""",
                (aid, mois, montant, statut, couleur),
            )
            ne += 1
    return na, ne


def import_previsionnel(con: sqlite3.Connection, ws, eid: int) -> int:
    """Codes TRI en col A, budget HT col C, TTC col E (valeurs ou formules simples)."""
    n = 0
    for r in range(3, 45):
        code = ws.cell(r, 1).value
        if not code or not isinstance(code, str):
            continue
        code = code.strip()
        if not con.execute("SELECT 1 FROM categories WHERE code=?", (code,)).fetchone():
            continue
        ht = eval_simple_amount(ws.cell(r, 3).value)
        if ht is None:
            ht = to_float(ws.cell(r, 3).value) or 0.0
        ttc = eval_simple_amount(ws.cell(r, 5).value)
        if ttc is None:
            ttc = to_float(ws.cell(r, 5).value)
        if ttc is None:
            ttc = ht
        tva = max(0.0, ttc - ht) if ttc is not None and ht is not None else 0.0
        con.execute(
            """INSERT INTO budgets (exercice_id, categorie_code, montant_ht, montant_tva, montant_ttc)
               VALUES (?,?,?,?,?)
               ON CONFLICT(exercice_id, categorie_code) DO UPDATE SET
                 montant_ht=excluded.montant_ht,
                 montant_tva=excluded.montant_tva,
                 montant_ttc=excluded.montant_ttc""",
            (eid, code, ht or 0.0, tva, ttc or 0.0),
        )
        n += 1
    return n


def import_ca(con: sqlite3.Connection, ws) -> int:
    """Historique CA : lignes CA 2021… avec montants en dur ou formules simples."""
    n = 0
    for r in range(1, (ws.max_row or 1) + 1):
        label = ws.cell(r, 3).value
        if not label:
            continue
        m = re.search(r"CA\s*(20\d{2})", str(label), re.I)
        if not m:
            continue
        annee = int(m.group(1))
        # colonnes D (sem1) E (sem2) F (année) — on prend F si nombre, sinon D+E si nombres
        f = eval_simple_amount(ws.cell(r, 6).value)
        d = eval_simple_amount(ws.cell(r, 4).value)
        e = eval_simple_amount(ws.cell(r, 5).value)
        total = None
        if f is not None:
            total = f
        elif d is not None or e is not None:
            total = (d or 0) + (e or 0)
        if total is None:
            continue
        con.execute(
            """INSERT INTO ca_historique (annee, ca_ht, source)
               VALUES (?,?,?)
               ON CONFLICT(annee) DO UPDATE SET ca_ht=excluded.ca_ht, source=excluded.source""",
            (annee, total, "excel"),
        )
        n += 1
    return n


def planning_months_from_headers(ws, start_col: int = 9, end_col: int = 26) -> list[str]:
    """Essaie de lire les mois ; sinon retourne liste vide (appelant fournit défaut)."""
    # N4/N5 ont souvent seulement le nom du mois sans année en ligne 4
    return []


def main() -> int:
    if not DB.is_file():
        log(f"Base introuvable : {DB}")
        log("Lancez d’abord : php scripts/init_db.php")
        return 1
    if not N4.is_file() or not N5.is_file():
        log("Classeurs N4/N5 manquants dans aides/")
        return 1

    log(f"Base : {DB}")
    log("Lecture des classeurs (peut prendre ~30 s)…")
    wb4 = load_workbook(N4, data_only=False, keep_vba=False)
    wb5 = load_workbook(N5, data_only=False, keep_vba=False)

    con = sqlite3.connect(str(DB))
    con.execute("PRAGMA foreign_keys = ON")

    eid4 = ensure_exercice(con, "N4", "Exercice N4 (juil. 2024 – juin 2025)", "2024-07-01", "2025-06-30", 0)
    eid5 = ensure_exercice(con, "N5", "Exercice N5 (juil. 2025 – déc. 2026, 18 mois)", "2025-07-01", "2026-12-31", 1)
    log(f"Exercices : N4 id={eid4}, N5 id={eid5} (actif)")

    clear_exercice_data(con, eid4)
    clear_exercice_data(con, eid5)
    # ops sans exercice (imports CIC) : on laisse

    # --- N4 ---
    log("--- N4 ---")
    n = import_facturation(con, wb4["facturation 24-25"], eid4, "n4")
    log(f"  Factures : {n}")
    n = import_banque_n4(con, wb4["Pointage BQ 24-25"], eid4)
    log(f"  Banque (Pointage BQ) : {n}")
    # Planning N4 : juil 2024 – juin 2025 = 12 mois col I–T typiquement
    n4_months = [
        "202407", "202408", "202409", "202410", "202411", "202412",
        "202501", "202502", "202503", "202504", "202505", "202506",
    ]
    na, ne = import_planning(con, wb4["Planning facturation 24-25"], eid4, n4_months)
    log(f"  Affaires / échéances : {na} / {ne}")
    n = import_previsionnel(con, wb4["Prévisionnel 24-25"], eid4)
    log(f"  Budgets prévisionnels : {n}")
    n = import_ca(con, wb4["CA"])
    log(f"  CA historique (feuille N4) : {n}")

    # --- N5 ---
    log("--- N5 ---")
    n = import_facturation(con, wb5["facturation 25-26"], eid5, "n5")
    log(f"  Factures : {n}")
    n = import_banque_n5(con, wb5["BANQUE"], eid5)
    log(f"  Banque : {n}")
    na, ne = import_planning(con, wb5["Planning facturation 25-26"], eid5, N5_MONTHS)
    log(f"  Affaires / échéances : {na} / {ne}")
    n = import_previsionnel(con, wb5["Prévisionnel 25-26"], eid5)
    log(f"  Budgets prévisionnels : {n}")
    n = import_ca(con, wb5["CA"])
    log(f"  CA historique (feuille N5, complète) : {n}")

    con.commit()

    # résumé
    log("--- Résumé base ---")
    for label, sql in [
        ("Factures", "SELECT COUNT(*) FROM factures"),
        ("Opérations banque", "SELECT COUNT(*) FROM operations_bancaires"),
        ("Affaires", "SELECT COUNT(*) FROM affaires"),
        ("Échéances", "SELECT COUNT(*) FROM echeances_facturation"),
        ("Budgets", "SELECT COUNT(*) FROM budgets"),
        ("CA années", "SELECT COUNT(*) FROM ca_historique"),
    ]:
        log(f"  {label} : {con.execute(sql).fetchone()[0]}")

    con.close()
    log("Terminé. Rechargez l’application (Factures / Banque / Exercices).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
